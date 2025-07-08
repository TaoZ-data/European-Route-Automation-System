
# ========================================
# EUROPEAN ROUTE AUTOMATION SYSTEM
# Dynamic city discovery with configurable output
# ========================================


# Import all libraries
import requests
import pandas as pd
import folium
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
import json
import os
from datetime import datetime, timedelta
import time
from itertools import combinations


def calculate_arrival_time(departure_str, duration_str):
    """
    Calculate arrival time by adding duration to departure time.
    """
    departure = datetime.strptime(departure_str, "%H:%M")
    hours, minutes = 0, 0
    if "h" in duration_str:
        hours = int(duration_str.split("h")[0].strip())
        if "m" in duration_str:
            minutes = int(duration_str.split("h")[1].replace("m", "").strip())
    elif "m" in duration_str:
        minutes = int(duration_str.replace("m", "").strip())
    arrival = departure + timedelta(hours=hours, minutes=minutes)
    return arrival.strftime("%H:%M")

# ========================================
# 1. DYNAMIC CITY DISCOVERY
# ========================================

class DynamicCityDiscovery:
    """Automatically discover cities using APIs"""

    def __init__(self):
        self.overpass_url = "http://overpass-api.de/api/interpreter"
        self.geocoder = Nominatim(user_agent="route_automation_v2", timeout=10)
        self.city_cache = {}

    def discover_cities_by_country(self, country_name, min_population=50000, max_cities=20):
        """Discover cities using OpenStreetMap API"""

        print(f"🔍 Discovering cities in {country_name}...")

        # Primary query - cities with population data
        query = f"""
        [out:json][timeout:90];
        area["name"="{country_name}"]["admin_level"="2"]->.country;
        (
          node["place"="city"]["population"]["name"](area.country);
          node["place"="town"]["population"]["name"](area.country);
        );
        out center;
        """

        cities = []

        try:
            response = requests.post(self.overpass_url, data={'data': query}, timeout=120)
            data = response.json()

            for element in data.get('elements', []):
                tags = element.get('tags', {})

                if 'name' in tags:
                    # Extract population
                    population = 0
                    if 'population' in tags:
                        try:
                            population = int(tags['population'])
                        except (ValueError, TypeError):
                            population = 0

                    # Estimate if no population data
                    if population == 0:
                        place_type = tags.get('place', '')
                        if place_type == 'city':
                            population = 100000
                        elif place_type == 'town':
                            population = 50000

                    if population >= min_population:
                        city_info = {
                            'name': tags['name'],
                            'country': country_name,
                            'population': population,
                            'lat': element.get('lat'),
                            'lon': element.get('lon'),
                            'coordinates': (element.get('lat'), element.get('lon')),
                            'place_type': tags.get('place', 'city')
                        }

                        # Avoid duplicates
                        if not any(c['name'] == city_info['name'] for c in cities):
                            cities.append(city_info)

        except Exception as e:
            print(f"⚠️  API query failed for {country_name}: {e}")

        # If API fails, use geocoding fallback
        if len(cities) < 3:
            print(f"🔄 Using geocoding fallback for {country_name}")
            cities = self._geocoding_fallback(country_name, min_population, max_cities)

        # Sort and limit
        cities.sort(key=lambda x: x['population'], reverse=True)
        cities = cities[:max_cities]

        print(f"✅ Discovered {len(cities)} cities in {country_name}")

        # Cache results
        self.city_cache[country_name] = cities

        return cities

    def _geocoding_fallback(self, country_name, min_population, max_cities):
        """Fallback using geocoding for known major cities"""

        # Major cities by country (fallback only)
        major_cities_lookup = {
            'France': ['Paris', 'Lyon', 'Marseille', 'Toulouse', 'Nice', 'Nantes', 'Strasbourg', 'Montpellier'],
            'Germany': ['Berlin', 'Hamburg', 'Munich', 'Cologne', 'Frankfurt', 'Stuttgart', 'Düsseldorf', 'Dortmund'],
            'Spain': ['Madrid', 'Barcelona', 'Valencia', 'Seville', 'Bilbao', 'Malaga', 'Murcia', 'Las Palmas'],
            'Italy': ['Rome', 'Milan', 'Naples', 'Turin', 'Palermo', 'Genoa', 'Bologna', 'Florence'],
            'Netherlands': ['Amsterdam', 'Rotterdam', 'The Hague', 'Utrecht', 'Eindhoven', 'Tilburg'],
            'Belgium': ['Brussels', 'Antwerp', 'Ghent', 'Charleroi', 'Liège', 'Bruges'],
            'Poland': ['Warsaw', 'Krakow', 'Lodz', 'Wroclaw', 'Poznan', 'Gdansk'],
            'Czech Republic': ['Prague', 'Brno', 'Ostrava', 'Plzen'],
            'Austria': ['Vienna', 'Graz', 'Linz', 'Salzburg', 'Innsbruck'],
            'Portugal': ['Lisbon', 'Porto', 'Vila Nova de Gaia', 'Amadora'],
            'Switzerland': ['Zurich', 'Geneva', 'Basel', 'Lausanne', 'Bern'],
            'Sweden': ['Stockholm', 'Gothenburg', 'Malmo', 'Uppsala'],
            'Norway': ['Oslo', 'Bergen', 'Trondheim', 'Stavanger'],
            'Denmark': ['Copenhagen', 'Aarhus', 'Odense', 'Aalborg'],
        }

        cities = []
        city_names = major_cities_lookup.get(country_name, [])

        for city_name in city_names[:max_cities]:
            try:
                query = f"{city_name}, {country_name}"
                location = self.geocoder.geocode(query, timeout=15)

                if location:
                    city_info = {
                        'name': city_name,
                        'country': country_name,
                        'population': min_population,
                        'lat': location.latitude,
                        'lon': location.longitude,
                        'coordinates': (location.latitude, location.longitude),
                        'place_type': 'city'
                    }
                    cities.append(city_info)
                    time.sleep(2)

            except Exception as e:
                print(f"⚠️  Geocoding failed for {city_name}: {e}")
                continue

        return cities

# ========================================
# 2. ROUTE GENERATOR
# ========================================

class IntelligentRouteGenerator:
    """Generate routes from discovered cities"""

    def __init__(self):
        self.city_discovery = DynamicCityDiscovery()

    def generate_routes_from_config(self, config):
        """Generate routes based on configuration"""

        print("🤖 INTELLIGENT ROUTE GENERATION")
        print("=" * 50)

        routes = []

        if 'countries' in config:
            countries = config['countries']
            min_pop = config.get('min_population', 100000)

            all_cities = {}
            for country in countries:
                cities = self.city_discovery.discover_cities_by_country(country, min_pop, 20)
                for city in cities:
                    city_key = f"{city['name']}_{country}"
                    all_cities[city_key] = city

        elif 'auto_discover_europe' in config:
            countries = self.city_discovery.discover_european_countries()
            countries = countries[:config.get('max_countries', 4)]

            all_cities = {}
            min_pop = config.get('min_population', 100000)

            for country in countries:
                cities = self.city_discovery.discover_cities_by_country(country, min_pop, 15)
                for city in cities:
                    city_key = f"{city['name']}_{country}"
                    all_cities[city_key] = city

        else:
            print("❌ No valid discovery configuration found")
            return []

        print(f"\n🚌 Generating routes from {len(all_cities)} discovered cities...")

        cities_list = list(all_cities.values())
        routes = self._generate_optimal_routes(cities_list, config)

        print(f"✅ Generated {len(routes)} routes")
        return routes

    def _generate_optimal_routes(self, cities, config):
        """Generate optimal routes from cities"""

        routes = []
        route_id = 1

        min_distance = config.get('min_distance_km', 100)
        max_distance = config.get('max_distance_km', 800)
        max_routes = config.get('max_routes', 50)

        for i, city1 in enumerate(cities):
            for city2 in cities[i+1:]:

                distance = geodesic(city1['coordinates'], city2['coordinates']).kilometers

                if min_distance <= distance <= max_distance:
                    route = {
                        'id': f'EUR_{route_id:03d}',
                        'origin': city1['name'],
                        'destination': city2['name'],
                        'origin_country': city1['country'],
                        'dest_country': city2['country'],
                        'estimated_distance': round(distance, 1),
                        'route_type': 'domestic' if city1['country'] == city2['country'] else 'international'
                    }
                    routes.append(route)
                    route_id += 1

                    if len(routes) >= max_routes:
                        return routes

        return routes

# ========================================
# 3. DATA COLLECTOR
# ========================================

class DynamicDataCollector:
    """Collect route data using discovered cities"""

    def __init__(self, city_discovery):
        self.city_discovery = city_discovery

    def get_route_info(self, origin_name, destination_name, discovered_cities):
        """Get route information from discovered city data"""

        origin_city = self._find_city_in_discovered(origin_name, discovered_cities)
        dest_city = self._find_city_in_discovered(destination_name, discovered_cities)

        if not origin_city or not dest_city:
            return None

        origin_coords = origin_city['coordinates']
        dest_coords = dest_city['coordinates']

        distance = geodesic(origin_coords, dest_coords).kilometers
        travel_time_minutes = int(distance / 60 * 60)

        route_info = {
            'origin': origin_name,
            'destination': destination_name,
            'origin_coords': origin_coords,
            'dest_coords': dest_coords,
            'distance_km': round(distance, 1),
            'travel_time_minutes': travel_time_minutes,
            'travel_time_formatted': f"{travel_time_minutes//60}h {travel_time_minutes%60}m"
        }

        return route_info

    def _find_city_in_discovered(self, city_name, discovered_cities):
        """Find city in discovered data"""

        for city_key, city_data in discovered_cities.items():
            if city_data['name'].lower() == city_name.lower():
                return city_data

        return None

# ========================================
# 4. MAIN AUTOMATION SYSTEM
# ========================================

class CompleteDynamicAutomation:
    """Main automation system"""

    def __init__(self):
        self.route_generator = IntelligentRouteGenerator()

        os.makedirs('outputs/maps', exist_ok=True)
        os.makedirs('outputs/timetables', exist_ok=True)
        os.makedirs('outputs/pricing', exist_ok=True)
        os.makedirs('outputs/reports', exist_ok=True)

    def run_automation(self, config):
        """Run complete automation system"""

        print("🚀 DYNAMIC EUROPEAN ROUTE AUTOMATION")
        print("=" * 60)
        print("🎯 Automated discovery and documentation generation")
        print("=" * 60)

        routes = self.route_generator.generate_routes_from_config(config)

        if not routes:
            print("❌ No routes generated")
            return None

        discovered_cities = self.route_generator.city_discovery.city_cache
        flat_cities = {}
        for country_cities in discovered_cities.values():
            for city in country_cities:
                flat_cities[f"{city['name']}_{city['country']}"] = city

        data_collector = DynamicDataCollector(self.route_generator.city_discovery)

        # Get configurable number of routes to process (default 3 for demo)
        process_count = config.get('process_routes', 3)
        demo_routes = routes[:process_count]

        print(f"\n📋 Processing {len(demo_routes)} routes for demonstration:")
        print(f"    (Total routes discovered: {len(routes)})")

        successful_count = 0
        routes_data = {}

        for i, route in enumerate(demo_routes, 1):
            route_id = route['id']
            origin = route['origin']
            destination = route['destination']

            print(f"\n🚌 Route {i}/{len(demo_routes)}: {route_id}")
            print(f"   📍 {origin} → {destination}")

            route_info = data_collector.get_route_info(origin, destination, flat_cities)

            if route_info:
                routes_data[route_id] = route_info
                successful_count += 1
                print(f"   ✅ Distance: {route_info['distance_km']} km, Travel: {route_info['travel_time_formatted']}")

                self._generate_route_documents(route_info, route_id)
            else:
                print(f"   ❌ Failed to process route data")

        if routes_data:
            self._generate_reports(routes_data, routes, discovered_cities, flat_cities)

        print(f"\n🎉 AUTOMATION COMPLETE!")
        print(f"📊 Routes processed: {successful_count}/{len(demo_routes)}")
        print(f"🌍 Countries discovered: {len(discovered_cities)}")
        print(f"🏙️  Cities discovered: {len(flat_cities)}")
        print(f"🚌 Total routes generated: {len(routes)}")
        print(f"📁 All files saved to outputs/ directory")

        return {
            'routes_processed': successful_count,
            'total_routes_generated': len(routes),
            'countries_discovered': len(discovered_cities),
            'cities_discovered': len(flat_cities),
            'routes_data': routes_data
        }

    def _generate_route_documents(self, route_info, route_id):
        """Generate documents for a route"""

        try:
            map_path = f"outputs/maps/{route_id}_route_map.html"
            self._create_route_map(route_info, route_id, map_path)

            timetable_path = f"outputs/timetables/{route_id}_timetable.xlsx"
            self._create_route_timetable(route_info, route_id, timetable_path)

            print(f"   📄 Documents: ✅ Map, ✅ Timetable")

        except Exception as e:
            print(f"   ❌ Document generation error: {e}")

    def _create_route_map(self, route_info, route_id, save_path):
        """Create interactive route map"""

        origin_coords = route_info['origin_coords']
        dest_coords = route_info['dest_coords']

        center_lat = (origin_coords[0] + dest_coords[0]) / 2
        center_lon = (origin_coords[1] + dest_coords[1]) / 2

        route_map = folium.Map(location=[center_lat, center_lon], zoom_start=6)

        folium.Marker(
            origin_coords,
            popup=f"<b>{route_info['origin']}</b><br>Origin<br>Route: {route_id}",
            icon=folium.Icon(color='green', icon='play')
        ).add_to(route_map)

        folium.Marker(
            dest_coords,
            popup=f"<b>{route_info['destination']}</b><br>Destination<br>Distance: {route_info['distance_km']} km",
            icon=folium.Icon(color='red', icon='stop')
        ).add_to(route_map)

        folium.PolyLine(
            locations=[origin_coords, dest_coords],
            weight=4,
            color='blue',
            popup=f"Route {route_id}: {route_info['distance_km']} km"
        ).add_to(route_map)

        route_map.save(save_path)

    def _create_route_timetable(self, route_info, route_id, ws):
        """
        Create professional timetable in the provided worksheet (ws)
        """
        from openpyxl.styles import Font

        ws.title = f"Route_{route_id}"

        ws['A1'] = f"EUROPEAN ROUTE TIMETABLE - {route_id}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = f"{route_info['origin']} → {route_info['destination']} | {route_info['distance_km']} km | {route_info['travel_time_formatted']}"

        headers = ['Service', 'Departure', 'Arrival', 'Duration', 'Capacity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)

        services = [
            ('Morning Express', '07:00', '2h 30m', 55),
            ('Standard Service', '12:00', '3h 15m', 48),
            ('Evening Express', '18:30', '2h 45m', 55)
        ]

        for row, (service, dept, duration, capacity) in enumerate(services, 5):
            arrival = calculate_arrival_time(dept, duration)
            ws.cell(row=row, column=1, value=service)
            ws.cell(row=row, column=2, value=dept)
            ws.cell(row=row, column=3, value=arrival)
            ws.cell(row=row, column=4, value=duration)
            ws.cell(row=row, column=5, value=capacity)

        for col in range(1, 6):
            ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = 18

    def generate_all_timetables(self, routes_data):
        from openpyxl import Workbook

        wb = Workbook()
        first = True
        for route_id, route_info in routes_data.items():
            if first:
                ws = wb.active
                ws.title = f"Route_{route_id}"
                first = False
            else:
                ws = wb.create_sheet(title=f"Route_{route_id}")
            self._create_route_timetable(route_info, route_id, ws)
        wb.save("outputs/timetables/all_timetables.xlsx")
        print("✅ All timetables saved to outputs/timetables/all_timetables.xlsx")

    def _generate_reports(self, routes_data, all_routes, discovered_countries, discovered_cities):
        """Generate comprehensive reports"""

        wb = Workbook()
        ws = wb.active
        ws.title = "Route_Analysis"

        ws['A1'] = "EUROPEAN ROUTE AUTOMATION - COMPREHENSIVE ANALYSIS"
        ws['A1'].font = Font(bold=True, size=16)

        ws['A3'] = "DISCOVERY STATISTICS:"
        ws['A3'].font = Font(bold=True)

        ws['A4'] = f"Countries analyzed: {len(discovered_countries)}"
        ws['A5'] = f"Cities discovered: {len(discovered_cities)}"
        ws['A6'] = f"Routes generated: {len(all_routes)}"
        ws['A7'] = f"Routes processed: {len(routes_data)}"
        ws['A8'] = f"Automation date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        ws['A10'] = "PROCESSED ROUTES:"
        ws['A10'].font = Font(bold=True)

        headers = ['Route ID', 'Origin', 'Destination', 'Distance (km)', 'Travel Time', 'Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = Font(bold=True)

        row = 13
        for route_id, route_info in routes_data.items():
            ws.cell(row=row, column=1, value=route_id)
            ws.cell(row=row, column=2, value=route_info['origin'])
            ws.cell(row=row, column=3, value=route_info['destination'])
            ws.cell(row=row, column=4, value=route_info['distance_km'])
            ws.cell(row=row, column=5, value=route_info['travel_time_formatted'])
            ws.cell(row=row, column=6, value='European')
            row += 1

        for col in range(1, 7):
            ws.column_dimensions[ws.cell(row=12, column=col).column_letter].width = 20

        summary_path = "outputs/reports/european_route_analysis.xlsx"
        wb.save(summary_path)
        print(f"📊 Analysis report: {summary_path}")

        self._generate_pricing_analysis(routes_data)

    def _generate_pricing_analysis(self, routes_data):
        """Generate pricing analysis"""

        wb = Workbook()
        ws = wb.active
        ws.title = "Pricing_Analysis"

        ws['A1'] = "EUROPEAN ROUTES - COMPETITIVE PRICING ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        headers = ['Route', 'Distance (km)', 'Base Price (€)', 'Premium Price (€)', 'Market Position']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)

        row = 5
        for route_id, route_info in routes_data.items():
            distance = route_info['distance_km']
            base_price = round(distance * 0.08 + 5, 2)
            premium_price = round(base_price * 1.3, 2)

            if distance < 200:
                market_position = "Regional"
            elif distance < 500:
                market_position = "National"
            else:
                market_position = "International"

            route_name = f"{route_info['origin']} → {route_info['destination']}"

            ws.cell(row=row, column=1, value=route_name)
            ws.cell(row=row, column=2, value=distance)
            ws.cell(row=row, column=3, value=base_price)
            ws.cell(row=row, column=4, value=premium_price)
            ws.cell(row=row, column=5, value=market_position)
            row += 1

        for col in range(1, 6):
            ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = 22

        pricing_path = "outputs/pricing/european_pricing_analysis.xlsx"
        wb.save(pricing_path)
        print(f"💰 Pricing analysis: {pricing_path}")

# ========================================
# 5. MAIN EXECUTION
# ========================================

def run_route_automation(countries=['France', 'Spain'], num_routes=3, min_population=150000):
    """
    Run European route automation with customizable parameters

    Parameters:
    - countries: List of countries to analyze (default: ['France', 'Spain'])
    - num_routes: Number of routes to process (default: 3)
    - min_population: Minimum city population (default: 150000)
    """

    print("🚀 EUROPEAN ROUTE AUTOMATION SYSTEM")
    print("🌍 Dynamic city discovery with configurable output")
    print("=" * 60)

    # Create configuration from parameters
    config = {
        'countries': countries,
        'min_population': min_population,
        'max_routes': num_routes * 10,  # Discover more routes than we process
        'process_routes': num_routes,
        'min_distance_km': 150,
        'max_distance_km': 700
    }

    print(f"📋 Configuration:")
    print(f"   • Countries: {countries}")
    print(f"   • Routes to process: {num_routes}")
    print(f"   • Min population: {min_population:,}")
    print(f"   • Max routes to discover: {config['max_routes']}")

    # Run automation
    automation = CompleteDynamicAutomation()
    results = automation.run_automation(config)

    if results:
        print(f"\n✨ AUTOMATION COMPLETE!")
        print(f"🎉 Successfully processed {results['routes_processed']} routes")
        print(f"📊 Total routes generated: {results['total_routes_generated']}")
        print(f"🌍 Countries analyzed: {results['countries_discovered']}")
        print(f"🏙️  Cities discovered: {results['cities_discovered']}")
        print(f"📁 All documentation saved to outputs/ directory")
        print("=" * 60)
    else:
        print("\n❌ Automation failed. Check API connectivity and try again.")

    return results

if __name__ == "__main__":
    # Default execution - France, Spain, 3 routes
    results = run_route_automation()
    automation = CompleteDynamicAutomation()
    automation.generate_all_timetables(results['routes_data'])


    # Different countries and routes:
    # results = run_route_automation(['France', 'Germany', 'Spain'], 10)
    # automation = CompleteDynamicAutomation()
    # automation.generate_all_timetables(results['routes_data'])

    # Single country:
    # results = run_route_automation(['Netherlands'], 2)
    # automation = CompleteDynamicAutomation()
    # automation.generate_all_timetables(results['routes_data'])

    # Larger population threshold:
    # results = run_route_automation(['France', 'Germany'], 3, 200000)
    # automation = CompleteDynamicAutomation()
    # automation.generate_all_timetables(results['routes_data'])

